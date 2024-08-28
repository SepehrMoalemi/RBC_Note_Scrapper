# Purpose:
# %------------------------------------------ Packages -------------------------------------------% #
import time
import json
import re
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from bs4      import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by  import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support    import expected_conditions as EC

# %------------------------------------------- Classes --------------------------------------------% #

class WebScraper():
    def __init__(self, DOMAIN, SITE, 
                 MAX_LOADING_TIME=60,
                 MAX_NUMBER_FUNDS=20)-> None:
        # Set const web parameters
        self.DOMAIN = DOMAIN
        self.SITE   = SITE
        
        self.MAX_LOADING_TIME = MAX_LOADING_TIME
        
        # Construct full url
        self.URL = self.DOMAIN + self.SITE
        self.contrustURL = lambda site: self.DOMAIN + site
        
        # Find table of notes on the website
        self.table = self.getTableOfProductsFromURL()
        
        # Get a dictionary of fund name and their hyperlinks, issue and maturity date
        self.funds_directory = self.getFundNameAndHyperlink(MAX_NUMBER_FUNDS=MAX_NUMBER_FUNDS)
        
    # %---------------------------------------- Methods --------------------------------------% #
    # Purpose: Get Table of notes from URL
    """ 
        !(specific to RBC Notes website)
    """
    def getTableOfProductsFromURL(self):
        # Create a new instance of the Chrome driver
        driver = webdriver.Chrome()
    
        # Load Page and wait till notes table is loaded
        condition = EC.presence_of_all_elements_located((By.CLASS_NAME, 'note-row'))
        driver = self.waitTillPageLoads(self.URL, driver, condition)
        
        # Get the page source after JavaScript has executed
        page_source = driver.page_source

        # Parse the page source with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Get Table Elements
        table = soup.find_all('tr', class_='note-row')
        
        # If table is empty, print error message
        if len(table) == 0:
            print(f'Table NOT Found at {self.URL}.')
            
        # Close the browser
        driver.quit()

        return table
    
    # Purpose: Get a dictionary of fund name and hyperlinks given table of notes
    def getFundNameAndHyperlink(self, MAX_NUMBER_FUNDS=20):
        funds = {}
        for row in self.table:
            # Break if we met max number of funds
            if len(funds) >= MAX_NUMBER_FUNDS:
                break
            
            # Find fund name and hyperlink
            name_element = row.find('td', align='left').find('a')
            fund_name = name_element.text.strip()
            relHyperlink = name_element['href']
            hyperlink = self.contrustURL(relHyperlink)
            
            # Find issue and maturity date
            issue_date    = row.find_all('td')[9].text.strip()
            maturity_date = row.find_all('td')[10].text.strip()
            
            # Add to dictionary
            funds[fund_name] = {'hyperlink':     hyperlink,
                                'issue_date':    issue_date,
                                'maturity_date': maturity_date}
        return funds
    
    # Purpose: Set Reference Information from fund page
    def setRefInfoFromFundPage(self):
        for fund, fund_data in self.funds_directory.items():
            # Get URL from hyperlink
            url = fund_data['hyperlink']
            
            # Create a new instance of the Chrome driver
            driver = webdriver.Chrome()
            
            # Load Page and wait till reference table is loaded
            condition = EC.presence_of_all_elements_located((By.ID, 'divRefInfo'))
            driver = self.waitTillPageLoads(url, driver, condition)
            
            # Get the page source after JavaScript has executed
            page_source = driver.page_source

            # Parse the page source with BeautifulSoup
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Find the div with id "divRefInfo"
            ref_info_div = soup.find('div', id='divRefInfo')

            # Find the table within the div
            table = ref_info_div.find('table', class_='table')

            # Create an empty dictionary to store the data
            refInfo = {}

            # Loop through the rows of the table and extract data
            for row in table.find_all('tr'):
                cells = row.find_all('td')
                if len(cells) == 2:
                    key = cells[0].text.strip()
                    value = cells[1].text.strip()
                    refInfo[key] = value
            
            # Append reference information to dictionary
            self.funds_directory[fund]['refInfo'] = refInfo
            
    # Purpose: Categories the funds
    def categorizeFunds(self):
        # Dictionary to store the categories
        fund_categories = {}
        
        # Iterate through funds and categorize them
        for fund in self.funds_directory:
            # Remove "F-Class"
            trimmed_name = fund.replace("F-Class", "").strip()
            
            # Remove "Series number"
            trimmed_name = re.sub(r'Series \d+', '', trimmed_name).strip()
            
            # Remove any ","
            trimmed_name = trimmed_name.replace(",", "").strip()
            
            # Remove any numbers before the % sign including the % and "Securities"
            trimmed_name = re.sub(r'\d+(\.\d+)?% Securities', '', trimmed_name).strip()

            # Check if this fund is already in the dictionary
            if trimmed_name not in fund_categories:
                fund_categories[trimmed_name] = []
            
            # Append fund to the category
            fund_categories[trimmed_name].append(fund)
        return fund_categories
    
    # Purpose: Export funds category to csv
    def exportFundsCategoryToCSV(self, fund_categories):
        # Create a new workbook
        wb = Workbook()

        # Iterate through each category
        for i, (category, funds) in enumerate(fund_categories.items()):
            # Create a new sheet for the category
            ws = wb.create_sheet(title=str(i + 1))

            # Merge cells for category name
            end_column = 4 + len(self.funds_directory[funds[0]]["refInfo"])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
            ws.cell(row=1, column=1, value=category)
            ws.cell(row=1, column=1).font = Font(bold=True)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # Set column titles
            ws.append(["Fund name"] + list(self.funds_directory[funds[0]]["refInfo"].keys()) + ["F-Class", "Issued Date", "Maturity Date"])

            # Populate the data
            for fund in funds:
                info = self.funds_directory[fund]
                fund_name = fund
                ref_info = info["refInfo"]
                f_class = "F-Class" if "F-Class" in fund else ""
                issued_date = info["issue_date"]
                maturity_date = info["maturity_date"]
                ws.append([fund_name] + [ref_info.get(key, "") for key in ref_info] + [f_class, issued_date, maturity_date])

        # Remove the default sheet created and save the workbook
        del wb["Sheet"]
        wb.save("funds_data.xlsx")

    # Purpose: Wait till page loads
    def waitTillPageLoads(self, url, driver, condition):
        # Start Timer to load page
        start_time = time.time()
        
        # Open the URL in the browser
        driver.get(url)
            
        # Wait for the page to load note table
        wait = WebDriverWait(driver, self.MAX_LOADING_TIME)
        wait.until(condition)
        
        # Record the end time
        end_time = time.time()
        loading_duration = end_time - start_time
        print(f'Page took {loading_duration:.2f} seconds to load.')
        
        return driver
    
    # Purpose: Prinf dictionary of funds
    def printFundsDirectory(self):
        print(json.dumps(self.funds_directory, indent=4))
        
# %-------------------------------------------- Main ---------------------------------------------% #
def main():
    # Set domain
    # URL of Solactive Canada Bank 40
    DOMAIN   = "https://www.rbcnotes.com"
    SITE_URL = "/Products?q=Solactive%20Canada%20Bank%2040"
    
    # Construct the Webscraper object
    data = WebScraper(DOMAIN, SITE_URL, 
                      MAX_NUMBER_FUNDS=20)    #! <--------- Change this to get more funds
    
    # Get the table elements from the URL
    print(len(data.funds_directory))
    
    # Set Reference Information from fund page
    data.setRefInfoFromFundPage()
    
    # Categorize the funds
    fund_categories = data.categorizeFunds()
    print(json.dumps(fund_categories, indent=4))
    
    data.printFundsDirectory()
    data.exportFundsCategoryToCSV(fund_categories)
# %--------------------------------------------- Run ---------------------------------------------% #
if __name__ == '__main__':
    print(f'{"Start":-^{50}}')
    main()
    print(f'{"End":-^{50}}')